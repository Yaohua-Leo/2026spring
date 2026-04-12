from __future__ import annotations

import argparse
import html
import json
import re
from dataclasses import asdict, dataclass
from datetime import time
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MARKDOWN_PATH = ROOT / "ICCM 2026 International Conference on Celestial Mechanics.md"
SCHEDULE_PATH = ROOT / "Schedule(可改)(1).xlsx"
CSS_PATH = ROOT / "templates" / "handbook.css"
TEMPLATE_PATH = ROOT / "templates" / "handbook_template.html"

DEFAULT_HTML = ROOT / "output" / "pdf" / "iccm_2026_handbook.html"
DEFAULT_JSON = ROOT / "output" / "pdf" / "iccm_2026_handbook_data.json"

DAY_ORDER = [
    "Monday, April 20",
    "Tuesday, April 21",
    "Wednesday, April 22",
    "Thursday, April 23",
    "Friday, April 24",
]

PRACTICAL_INFO = {
    "venue": {
        "name": "SUSTech International Center for Mathematics Lecture Hall",
        "address": "Second Floor, Taizhou Building, Southern University of Science and Technology, 1088 Xueyuan Avenue, Nanshan District, Shenzhen.",
        "contact": "icm@sustech.edu.cn",
    },
    "hotels": [
        {
            "name": "Shenzhen Jinfeng Hotel",
            "address": "4168-B Liuxian Avenue, Nanshan District, Shenzhen.",
            "note": "Recommended in the reference conference guide stored in this workspace.",
        },
        {
            "name": "Yayuan Tanglang Hotel",
            "address": "Tianliao Building, Xueyuan Avenue, Xili Town, Nanshan District, Shenzhen.",
            "note": "Recommended in the reference conference guide stored in this workspace.",
        },
    ],
    "notes": [
        "Campus map and hotel suggestions are adapted from the reference PDF already present in this workspace.",
        "Wednesday afternoon is reserved for free discussions and collaboration time.",
        "Dinner is marked for Monday and Thursday in the schedule workbook.",
    ],
}

SICM_INTRO_PARAGRAPHS = [
    "The Shenzhen International Center for Mathematics, officially established on February 22, 2019, is a scientific research center funded by the Shenzhen Municipal Government and located in Southern University of Science and Technology. The scientific director is Fields Medalist Efim Zelmanov.",
    "It aims to advance scientific research in pure mathematics, applied and computational mathematics in Shenzhen, the Greater Bay Area, and beyond.",
    "The center will focus on important research fields in both pure and applied mathematics.",
    "It’s primarily goal is to promote the international scientific cooperation and advance research in mathematics and interdisciplinary fields of life science, information science, engineering, finance, etc.",
    "The SICM works in accordance with international standards for major international mathematics centers. It provides first-class research environments for researchers and visitors.",
    "Hosting annual thematic programs, conferences, and workshops.",
]

CONFERENCE_GUIDE = {
    "registration": [
        "April 20, 9:00-9:30 AM",
        "SICM 240A, Taizhou Building Second Floor",
    ],
    "wifi": [
        "ICMPub",
        "Password: ICM88888",
    ],
    "accommodation": [
        "Shenzhen Xili Navie S Hotel (Vanke Cloud City Branch), Building 1389, Liuxian Avenue, Nanshan District",
        "深圳西丽奈威S酒店（万科云城店）",
    ],
}

ASSET_PATHS = {
    "sustech_logo": ROOT / "assets" / "logos" / "sustech" / "sustech_combo3_english_horizontal.png",
    "sicm_logo": ROOT / "assets" / "logos" / "sicm-logo.png",
    "nsfc_logo": ROOT / "assets" / "logos" / "nsfc-eng-logo.jpg",
    "map_page": ROOT / "assets" / "maps" / "reference-map-page.png",
}


@dataclass
class Talk:
    date: str
    time: str
    title: str
    speaker: str
    affiliation: str
    abstract: str


@dataclass
class ScheduleEntry:
    day: str
    slot: str
    kind: str
    label: str
    title: str = ""
    speaker: str = ""
    affiliation: str = ""


def normalize_line(value: str) -> str:
    return re.sub(r"[ \t]+", " ", value.replace("\xa0", " ").strip())


def normalize_block_text(value: str) -> str:
    value = html.unescape(value).replace("\r\n", "\n").replace("\xa0", " ")
    return "\n".join(normalize_line(line) for line in value.splitlines()).strip()


def normalize_slot(value: object) -> str:
    if isinstance(value, time):
        return value.strftime("%H:%M")
    text = normalize_line(str(value))
    text = re.sub(r"\b(\d{1,2}):(\d{2})\b", lambda m: f"{int(m.group(1)):02d}:{m.group(2)}", text)
    text = re.sub(r"\s*-\s*", " - ", text)
    return text


def slot_key(value: str) -> str:
    return normalize_slot(value).replace(" - ", "-")


def shared_event_label(value: str) -> bool:
    compact = normalize_line(value).lower()
    return compact in {"coffee break", "lunch time", "dinner time", "no dinner"}


def schedule_day_label(raw: object) -> str:
    text = normalize_line(str(raw))
    match = re.fullmatch(r"([A-Za-z]+ \d{1,2}) \(([A-Za-z]+)\)", text)
    if match:
        return f"{match.group(2)}, {match.group(1)}"
    return text


def parse_markdown(path: Path) -> list[Talk]:
    text = path.read_text(encoding="utf-8")
    pattern = re.compile(
        r"(?ms)^(Monday|Tuesday|Wednesday|Thursday|Friday), .+?(?=^(?:Monday|Tuesday|Wednesday|Thursday|Friday), |\Z)"
    )
    talks: list[Talk] = []

    for match in pattern.finditer(text):
        block = normalize_block_text(match.group(0))
        if "\n\nAbstract\n\n" not in block:
            raise ValueError(f"Could not locate abstract separator in block:\n{block[:200]}")

        header, abstract = block.split("\n\nAbstract\n\n", 1)
        parts = [part.strip() for part in header.split("\n\n") if part.strip()]
        if len(parts) < 5:
            raise ValueError(f"Unexpected talk header structure:\n{header}")

        date, slot, title, speaker, affiliation = parts[:5]
        talks.append(
            Talk(
                date=date,
                time=normalize_slot(slot),
                title=title,
                speaker=speaker,
                affiliation=affiliation,
                abstract=normalize_block_text(abstract),
            )
        )

    return talks


def parse_schedule(path: Path, talks: Iterable[Talk]) -> list[ScheduleEntry]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    day_headers = {
        column: schedule_day_label(ws.cell(row=2, column=column).value)
        for column in range(2, ws.max_column + 1)
    }

    talk_lookup = {(talk.date, slot_key(talk.time)): talk for talk in talks}
    entries: list[ScheduleEntry] = []

    for row in range(3, ws.max_row + 1):
        slot = normalize_slot(ws.cell(row=row, column=1).value)
        row_values = {
            column: normalize_line(str(ws.cell(row=row, column=column).value))
            if ws.cell(row=row, column=column).value is not None
            else ""
            for column in range(2, ws.max_column + 1)
        }

        non_empty = [value for value in row_values.values() if value]
        if len(non_empty) == 1 and shared_event_label(non_empty[0]):
            row_values = {column: non_empty[0] for column in row_values}

        for column, day in day_headers.items():
            cell_text = row_values[column]
            talk = talk_lookup.get((day, slot_key(slot)))
            if talk:
                entries.append(
                    ScheduleEntry(
                        day=day,
                        slot=slot,
                        kind="talk",
                        label=talk.speaker,
                        title=talk.title,
                        speaker=talk.speaker,
                        affiliation=talk.affiliation,
                    )
                )
            elif cell_text:
                entries.append(
                    ScheduleEntry(
                        day=day,
                        slot=slot,
                        kind="event",
                        label=cell_text,
                    )
                )

    return entries


def render_inline(text: str) -> str:
    text = normalize_block_text(text).replace("\\[", "[").replace("\\]", "]")
    pieces: list[str] = []
    last_index = 0

    for match in re.finditer(r"\$[^$]+\$", text):
        pieces.append(html.escape(text[last_index:match.start()]))
        pieces.append(f'<span class="math">{match.group(0)}</span>')
        last_index = match.end()

    pieces.append(html.escape(text[last_index:]))
    return "".join(pieces)


def render_paragraphs(text: str) -> str:
    paragraphs = [part.strip() for part in re.split(r"\n\s*\n", text) if part.strip()]
    return "".join(
        f"<p>{render_inline(paragraph).replace(chr(10), '<br>')}</p>" for paragraph in paragraphs
    )


def path_uri(path: Path) -> str:
    return path.resolve().as_uri()


def with_page_number(section_html: str, page_number: int, show_number: bool = True) -> str:
    badge = f'<div class="page-number-badge">{page_number}</div>' if show_number else ""
    head, tail = section_html.rsplit("</section>", 1)
    return f"{head}{badge}</section>{tail}"


def page_cover(payload: dict) -> str:
    logos = payload["logos"]
    return f"""
    <section class="page cover">
      <svg class="orbit-svg" viewBox="0 0 1000 1400" aria-hidden="true">
        <g fill="none" stroke="rgba(255,255,255,0.16)" stroke-width="2">
          <ellipse cx="660" cy="360" rx="280" ry="140" />
          <ellipse cx="670" cy="390" rx="360" ry="210" transform="rotate(-18 670 390)" />
          <ellipse cx="610" cy="420" rx="410" ry="240" transform="rotate(26 610 420)" />
        </g>
        <g fill="rgba(240,217,166,0.88)">
          <circle cx="678" cy="408" r="10" />
          <circle cx="290" cy="1000" r="6" />
          <circle cx="840" cy="980" r="4" />
        </g>
      </svg>
      <div class="logo-band">
        <div class="logo-chip"><img src="{logos['sustech_logo']}" alt="SUSTech logo"></div>
        <div class="logo-chip sicm"><img src="{logos['sicm_logo']}" alt="SICM logo"></div>
        <div class="logo-chip"><img src="{logos['nsfc_logo']}" alt="NSFC logo"></div>
      </div>
      <div class="cover-body">
        <p class="cover-kicker">Conference Handbook</p>
        <h1 class="cover-title">ICCM 2026 <span class="line-two">International Conference on Celestial Mechanics</span></h1>
        <div class="cover-meta">
          <p class="cover-date">{payload['conference_dates']}</p>
        </div>
      </div>
      <div class="cover-footer">Shenzhen, China</div>
    </section>
    """


def page_sicm_intro(payload: dict) -> str:
    intro_paragraphs = "".join(f"<p>{html.escape(item)}</p>" for item in SICM_INTRO_PARAGRAPHS)
    return f"""
    <section class="page sicm-simple-page">
      <div class="page-shell">
        <div class="sicm-copy">
          {intro_paragraphs}
        </div>
        <div class="sicm-welcome">Welcome to the Shenzhen International Center for Mathematics!</div>
      </div>
    </section>
    """


def page_conference_guide() -> str:
    return f"""
    <section class="page guide-page">
      <div class="page-shell">
        <span class="page-kicker">Conference Guide</span>
        <h2 class="page-title">Conference Guide</h2>
        <p class="page-lead">Key on-site information for check-in, connectivity, and accommodation.</p>
        <div class="guide-grid">
          <div class="guide-card">
            <h3>Registration</h3>
            <p>{html.escape(CONFERENCE_GUIDE['registration'][0])}</p>
            <p>{html.escape(CONFERENCE_GUIDE['registration'][1])}</p>
          </div>
          <div class="guide-card">
            <h3>Wifi</h3>
            <p><strong>{html.escape(CONFERENCE_GUIDE['wifi'][0])}</strong></p>
            <p>{html.escape(CONFERENCE_GUIDE['wifi'][1])}</p>
          </div>
          <div class="guide-card accommodation">
            <h3>Accomodation</h3>
            <p>{html.escape(CONFERENCE_GUIDE['accommodation'][0])}</p>
            <p>{html.escape(CONFERENCE_GUIDE['accommodation'][1])}</p>
          </div>
        </div>
      </div>
    </section>
    """


def page_schedule_day(day: str, entries: list[ScheduleEntry]) -> str:
    timeline = []
    for entry in entries:
        if entry.kind == "talk":
            timeline.append(
                f"""
                <article class="timeline-entry talk">
                  <div class="timeline-time">{html.escape(entry.slot)}</div>
                  <div class="timeline-body no-title">
                    <div class="speaker">{render_inline(entry.speaker)}</div>
                    <div class="affiliation">{render_inline(entry.affiliation)}</div>
                  </div>
                </article>
                """
            )
        else:
            timeline.append(
                f"""
                <article class="timeline-entry event">
                  <div class="timeline-time">{html.escape(entry.slot)}</div>
                  <div class="timeline-body">
                    <div class="event-label">{render_inline(entry.label)}</div>
                  </div>
                </article>
                """
            )
    return f"""
    <section class="page schedule-day">
      <div class="page-shell">
        <span class="page-kicker">Schedule At A Glance</span>
        <div class="day-header">
          <div>
            <h2 class="page-title">{html.escape(day)}</h2>
          </div>
        </div>
        <div class="timeline">
          {''.join(timeline)}
        </div>
      </div>
    </section>
    """


def page_divider() -> str:
    return """
    <section class="page section-divider">
      <svg class="orbit-svg" viewBox="0 0 1000 1400" aria-hidden="true">
        <g fill="none" stroke="rgba(255,255,255,0.16)" stroke-width="2">
          <ellipse cx="650" cy="520" rx="360" ry="220" />
          <ellipse cx="620" cy="540" rx="290" ry="160" transform="rotate(22 620 540)" />
          <ellipse cx="630" cy="540" rx="430" ry="260" transform="rotate(-20 630 540)" />
        </g>
        <circle cx="644" cy="526" r="9" fill="rgba(244,216,155,0.92)" />
      </svg>
      <div class="divider-copy">
        <span class="page-kicker">Program Details</span>
        <h2>Titles &amp; Abstracts</h2>
        <p>Talks are presented in chronological order, with each abstract starting on a fresh page.</p>
      </div>
    </section>
    """


def page_talk(talk: Talk) -> str:
    is_dense = len(talk.abstract) > 1450 or len(talk.title) > 85
    section_class = "talk-section dense" if is_dense else "talk-section"
    return f"""
    <section class="{section_class}">
      <div class="talk-topline">
        <span class="pill">{html.escape(talk.date)}</span>
        <span class="pill time">{html.escape(talk.time)}</span>
      </div>
      <h2 class="talk-title">{render_inline(talk.title)}</h2>
      <div class="talk-meta">
        <div class="speaker">{render_inline(talk.speaker)}</div>
        <div class="affiliation">{render_inline(talk.affiliation)}</div>
      </div>
      <div class="abstract-block">
        <h3>Abstract</h3>
        {render_paragraphs(talk.abstract)}
      </div>
    </section>
    """


def page_notes(index: int) -> str:
    return f"""
    <section class="notes-page">
      <div class="notes-title">
        <h2>Notes</h2>
        <span>Page {index}</span>
      </div>
      <div class="notes-lines"></div>
      <div class="notes-footer">Questions, ideas, and follow-up threads.</div>
    </section>
    """


def page_back_cover(payload: dict) -> str:
    logos = payload["logos"]
    return f"""
    <section class="back-cover">
      <svg class="orbit-svg" viewBox="0 0 1000 1400" aria-hidden="true">
        <g fill="none" stroke="rgba(255,255,255,0.12)" stroke-width="2">
          <ellipse cx="280" cy="320" rx="250" ry="130" />
          <ellipse cx="300" cy="380" rx="330" ry="200" transform="rotate(-24 300 380)" />
        </g>
      </svg>
      <div class="back-inner">
        <span class="page-kicker">Closing Page</span>
        <h2>See you in Shenzhen.</h2>
        <p><strong>Contact:</strong> {html.escape(payload['contact_email'])}</p>
        <div class="mini-logos">
          <div class="logo-chip"><img src="{logos['sustech_logo']}" alt="SUSTech logo"></div>
          <div class="logo-chip sicm"><img src="{logos['sicm_logo']}" alt="SICM logo"></div>
        </div>
      </div>
    </section>
    """


def build_document(talks: list[Talk], schedule: list[ScheduleEntry], notes_pages: int) -> tuple[str, dict]:
    grouped_schedule: dict[str, list[ScheduleEntry]] = {day: [] for day in DAY_ORDER}
    for entry in schedule:
        grouped_schedule.setdefault(entry.day, []).append(entry)

    extra_dinners = {
        "Tuesday, April 21": "Dinner Time",
        "Friday, April 24": "Dinner Time",
    }
    for day, label in extra_dinners.items():
        if not any(item.slot == "17:30" for item in grouped_schedule.get(day, [])):
            grouped_schedule.setdefault(day, []).append(
                ScheduleEntry(day=day, slot="17:30", kind="event", label=label)
            )

    payload = {
        "conference_dates": "April 20-24, 2026",
        "talk_count": len(talks),
        "day_count": len(DAY_ORDER),
        "venue_name": PRACTICAL_INFO["venue"]["name"],
        "contact_email": PRACTICAL_INFO["venue"]["contact"],
        "practical_info": PRACTICAL_INFO,
        "logos": {key: path_uri(value) for key, value in ASSET_PATHS.items()},
        "talks": [asdict(talk) for talk in talks],
        "schedule": [asdict(entry) for entry in schedule],
    }

    sections = [
        page_cover(payload),
        page_sicm_intro(payload),
        page_conference_guide(),
    ]
    sections.extend(page_schedule_day(day, grouped_schedule[day]) for day in DAY_ORDER)
    sections.append(page_divider())
    sections.extend(page_talk(talk) for talk in talks)
    sections.append(page_back_cover(payload))

    numbered_sections = []
    for index, section in enumerate(sections, start=1):
        if index == 1 or index == len(sections):
            numbered_sections.append(with_page_number(section, index, show_number=False))
        else:
            numbered_sections.append(with_page_number(section, index - 1, show_number=True))
    sections = numbered_sections

    styles = CSS_PATH.read_text(encoding="utf-8")
    template = TEMPLATE_PATH.read_text(encoding="utf-8")
    document = template.replace("{{ STYLES }}", styles).replace("{{ CONTENT }}", "\n".join(sections))
    return document, payload


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate the ICCM 2026 conference handbook HTML.")
    parser.add_argument("--output-html", type=Path, default=DEFAULT_HTML)
    parser.add_argument("--output-json", type=Path, default=DEFAULT_JSON)
    parser.add_argument("--notes-pages", type=int, default=4)
    args = parser.parse_args()

    talks = parse_markdown(MARKDOWN_PATH)
    schedule = parse_schedule(SCHEDULE_PATH, talks)
    document, payload = build_document(talks, schedule, notes_pages=args.notes_pages)

    args.output_html.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_html.write_text(document, encoding="utf-8")
    args.output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Wrote {args.output_html}")
    print(f"Wrote {args.output_json}")


if __name__ == "__main__":
    main()
